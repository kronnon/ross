#!/usr/bin/env python3
"""
洛氏霍克交易系统回测 - 2025年BTC 5分钟数据
"""
import pymongo
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import numpy as np

# MongoDB连接
client = pymongo.MongoClient(os.getenv('MONGO_URL', os.getenv('MONGO_URI', 'mongodb://localhost:27017/')))
db = client['trading-data']

# 策略参数 - 放宽条件
CONFIG = {
    'trade_amount': 100,        # 交易金额 USDT
    'rsi_oversold': 35,         # RSI超卖 (原30 -> 35)
    'rsi_overbought': 65,       # RSI超买 (原70 -> 65)
    'ema_fast': 9,              # 快速EMA
    'ema_slow': 21,             # 慢速EMA
    'stop_loss_pct': 0.5,       # 止损比例
    'take_profit_pct': 1.5,     # 止盈比例
    'max_hold_bars': 24,        # 最大持仓K线数
    'min_trade_interval': 5,    # 最小交易间隔
    'min_signals': 1,           # 最少信号数 (原2 -> 1)
}

def calculate_ema(prices, period):
    """计算EMA"""
    ema = []
    multiplier = 2 / (period + 1)
    for i, price in enumerate(prices):
        if i < period - 1:
            ema.append(None)
        elif i == period - 1:
            ema.append(sum(prices[:period]) / period)
        else:
            ema.append((price - ema[-1]) * multiplier + ema[-1])
    return ema

def calculate_volatility(prices, period=20):
    """计算波动率"""
    volatility = []
    for i in range(len(prices)):
        if i < period:
            volatility.append(None)
        else:
            returns = np.diff(prices[i-period:i]) / prices[i-period:i-1]
            volatility.append(np.std(returns))
    return volatility

def run_backtest(records):
    """运行回测"""
    trades = []
    position = None
    entry_price = 0
    entry_time = 0
    bars_since_entry = 0
    last_trade_bar = -999
    
    # 预处理
    prices = [r['price'] for r in records]
    ema_fast = calculate_ema(prices, CONFIG['ema_fast'])
    ema_slow = calculate_ema(prices, CONFIG['ema_slow'])
    volatility = calculate_volatility(prices)
    
    for i in range(1, len(records)):
        curr = records[i]
        prev = records[i-1]
        
        rsi = curr.get('rsi')
        k = curr.get('kdj', {}).get('k')
        d = curr.get('kdj', {}).get('d')
        j = curr.get('kdj', {}).get('j')
        boll = curr.get('boll', {})
        
        prev_rsi = prev.get('rsi')
        prev_k = prev.get('kdj', {}).get('k')
        prev_d = prev.get('kdj', {}).get('d')
        prev_boll = prev.get('boll', {})
        
        # 入场逻辑
        if position is None:
            if i - last_trade_bar < CONFIG['min_trade_interval']:
                continue
                
            if not all([rsi, k, d, boll.get('upper')]):
                continue
            
            # 趋势过滤
            current_ema_fast = ema_fast[i] if i < len(ema_fast) else None
            current_ema_slow = ema_slow[i] if i < len(ema_slow) else None
            trend_up = current_ema_fast and current_ema_slow and current_ema_fast > current_ema_slow
            
            # 波动率过滤
            current_vol = volatility[i] if i < len(volatility) and volatility[i] else 0.01
            if current_vol < 0.003 or current_vol > 0.05:
                continue
            
            signals = []
            
            # RSI超卖回升
            if prev_rsi and prev_rsi < CONFIG['rsi_oversold'] and rsi > CONFIG['rsi_oversold']:
                signals.append("RSI回升")
            
            # KDJ金叉
            if prev_k and prev_d and prev_k < prev_d and k > d:
                signals.append("KDJ金叉")
            
            # BOLL下轨支撑
            if prev_boll.get('lower') and prev['price'] < prev_boll['lower'] and curr['price'] > boll['lower']:
                signals.append("BOLL支撑")
            
            # KDJ超卖区金叉
            if k < 30 and prev_k and prev_d and prev_k < prev_d and k > d:
                signals.append("KDJ超卖金叉")
            
            # J值反转
            if j and prev.get('kdj', {}).get('j'):
                if prev.get('kdj', {}).get('j') < 0 and j > 0:
                    signals.append("J值反转")
            
            # 多信号共振
            if len(signals) >= CONFIG['min_signals'] and trend_up:
                position = 'long'
                entry_price = curr['price']
                entry_time = curr['time']
                bars_since_entry = 0
                last_trade_bar = i
        
        # 出场逻辑
        else:
            pnl_pct = (curr['price'] - entry_price) / entry_price * 100
            
            exit_signal = None
            exit_reason = ""
            
            # 止损
            if pnl_pct <= -CONFIG['stop_loss_pct']:
                exit_signal = True
                exit_reason = f"止损{abs(pnl_pct):.2f}%"
            
            # 止盈
            elif pnl_pct >= CONFIG['take_profit_pct']:
                exit_signal = True
                exit_reason = f"止盈{pnl_pct:.2f}%"
            
            # KDJ死叉
            elif prev_k and prev_d and prev_k > prev_d and k < d:
                exit_signal = True
                exit_reason = "KDJ死叉"
            
            # RSI超买
            elif rsi and rsi > CONFIG['rsi_overbought']:
                exit_signal = True
                exit_reason = f"RSI超买{rsi:.0f}"
            
            # 超时
            bars_since_entry += 1
            if bars_since_entry >= CONFIG['max_hold_bars']:
                exit_signal = True
                exit_reason = f"超时{bars_since_entry}根"
            
            if exit_signal:
                profit = (curr['price'] - entry_price) / entry_price * CONFIG['trade_amount']
                trades.append({
                    'entry_time': entry_time,
                    'entry_time_str': datetime.fromtimestamp(entry_time/1000).strftime('%Y-%m-%d %H:%M'),
                    'exit_time': curr['time'],
                    'exit_time_str': datetime.fromtimestamp(curr['time']/1000).strftime('%Y-%m-%d %H:%M'),
                    'entry_price': entry_price,
                    'exit_price': curr['price'],
                    'hold_bars': bars_since_entry,
                    'exit_reason': exit_reason,
                    'profit_usd': profit,
                    'pnl_pct': pnl_pct
                })
                position = None
    
    # 最后未平仓
    if position == 'long':
        last = records[-1]
        profit = (last['price'] - entry_price) / entry_price * CONFIG['trade_amount']
        pnl_pct = (last['price'] - entry_price) / entry_price * 100
        trades.append({
            'entry_time': entry_time,
            'entry_time_str': datetime.fromtimestamp(entry_time/1000).strftime('%Y-%m-%d %H:%M'),
            'exit_time': last['time'],
            'exit_time_str': datetime.fromtimestamp(last['time']/1000).strftime('%Y-%m-%d %H:%M'),
            'entry_price': entry_price,
            'exit_price': last['price'],
            'hold_bars': bars_since_entry,
            'exit_reason': '数据结束',
            'profit_usd': profit,
            'pnl_pct': pnl_pct
        })
    
    return trades

def export_to_excel(trades, filename):
    """导出交易记录到Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "交易记录"
    
    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    
    # 边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    headers = ['序号', '入场时间', '出场时间', '入场价格', '出场价格', '持仓K线数', '出场原因', '盈利(USDT)', '盈亏比例(%)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # 写入数据
    for row, trade in enumerate(trades, 2):
        ws.cell(row=row, column=1, value=row-1).border = thin_border
        ws.cell(row=row, column=2, value=trade['entry_time_str']).border = thin_border
        ws.cell(row=row, column=3, value=trade['exit_time_str']).border = thin_border
        ws.cell(row=row, column=4, value=trade['entry_price']).border = thin_border
        ws.cell(row=row, column=5, value=trade['exit_price']).border = thin_border
        ws.cell(row=row, column=6, value=trade['hold_bars']).border = thin_border
        ws.cell(row=row, column=7, value=trade['exit_reason']).border = thin_border
        
        profit_cell = ws.cell(row=row, column=8, value=round(trade['profit_usd'], 2))
        profit_cell.border = thin_border
        if trade['profit_usd'] > 0:
            profit_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif trade['profit_usd'] < 0:
            profit_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        pnl_cell = ws.cell(row=row, column=9, value=round(trade['pnl_pct'], 2))
        pnl_cell.border = thin_border
        if trade['pnl_pct'] > 0:
            pnl_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif trade['pnl_pct'] < 0:
            pnl_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 14
    
    wb.save(filename)
    print(f"Excel已保存: {filename}")

def main():
    print("=" * 60)
    print("洛氏霍克交易系统回测 - 2025年BTC 5分钟数据")
    print("=" * 60)
    
    # 获取数据
    collection = db['btc_5m']
    records = list(collection.find().sort('time', 1))
    print(f"\n加载数据: {len(records)} 条")
    
    # 运行回测
    print("\n运行回测...")
    trades = run_backtest(records)
    print(f"交易次数: {len(trades)}")
    
    # 统计
    winning_trades = [t for t in trades if t['profit_usd'] > 0]
    losing_trades = [t for t in trades if t['profit_usd'] < 0]
    total_profit = sum(t['profit_usd'] for t in trades)
    
    print(f"\n=== 回测统计 ===")
    print(f"总交易次数: {len(trades)}")
    print(f"盈利次数: {len(winning_trades)} ({len(winning_trades)/len(trades)*100:.1f}%)" if trades else "盈利次数: 0")
    print(f"亏损次数: {len(losing_trades)} ({len(losing_trades)/len(trades)*100:.1f}%)" if trades else "亏损次数: 0")
    print(f"总盈亏: {total_profit:.2f} USDT")
    
    if trades:
        avg_profit = total_profit / len(trades)
        print(f"平均盈亏: {avg_profit:.2f} USDT")
        print(f"最大盈利: {max(t['profit_usd'] for t in trades):.2f} USDT")
        print(f"最大亏损: {min(t['profit_usd'] for t in trades):.2f} USDT")
    
    # 出场原因统计
    print(f"\n=== 出场原因统计 ===")
    exit_reasons = {}
    for t in trades:
        reason = t['exit_reason']
        exit_reasons[reason] = exit_reasons.get(reason, 0) + 1
    for reason, count in sorted(exit_reasons.items(), key=lambda x: -x[1]):
        print(f"  {reason}: {count}次")
    
    # 导出Excel
    output_dir = os.path.expanduser("~/.openclaw/workspaces/ross/outputs")
    os.makedirs(output_dir, exist_ok=True)
    excel_file = os.path.join(output_dir, "rocky_hedgehog_trades_2025.xlsx")
    export_to_excel(trades, excel_file)

if __name__ == "__main__":
    main()
