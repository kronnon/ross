"""
回测引擎模块
"""

from typing import List, Optional, Dict
from datetime import datetime
import random
import os

from config import StrategyConfig
from signals import SignalGenerator, create_signal_generator
from risk import RiskManager, create_risk_manager
from position import PositionManager, Trade, create_position_manager


class BacktestEngine:
    """回测引擎"""
    
    def __init__(self, config: StrategyConfig):
        """
        初始化回测引擎
        
        Args:
            config: 策略配置（由量化系统传入）
        """
        if config is None:
            raise ValueError("config 必须由量化系统传入，不能为None")
        
        self.config = config
        
        # 初始化组件
        self.signal_generator = create_signal_generator(self.config)
        self.risk_manager = create_risk_manager(self.config)
        self.position_manager = create_position_manager(self.config, self.risk_manager)
        
        # 状态
        self.last_trade_bar = -self.config.min_trade_interval
        self.missed_signals = []
        
        # 多周期数据
        self.ht_data = None  # 大周期数据
    
    def load_higher_timeframe_data(self, ht_records: List[dict]):
        """
        加载大周期数据（用于多周期确认）
        
        Args:
            ht_records: 大周期K线数据列表
        """
        if ht_records:
            self.ht_data = {
                'closes': [r['close'] for r in ht_records],
                'highs': [r['high'] for r in ht_records],
                'lows': [r['low'] for r in ht_records],
            }
    
    def run(self, records: List[dict], ht_records: List[dict] = None) -> tuple[List[Trade], List[dict]]:
        """
        运行回测
        
        Args:
            records: K线数据列表
            ht_records: 大周期K线数据列表（可选，用于多周期确认）
        
        Returns:
            (交易列表, 错过的信号列表)
        """
        # 加载大周期数据
        if ht_records:
            self.load_higher_timeframe_data(ht_records)
        
        self._reset()
        
        # 提取数据
        closes = [r['close'] for r in records]
        opens = [r['open'] for r in records]
        highs = [r['high'] for r in records]
        lows = [r['low'] for r in records]
        times = [r['time'] for r in records]
        
        # 遍历K线
        for i in range(50, len(records)):
            current_price = closes[i]
            current_open = opens[i]
            current_high = highs[i]
            current_low = lows[i]
            current_time = times[i]
            
            # 更新持仓状态
            self.position_manager.update_positions(current_price)
            
            # 检查出场
            self._check_exits(i, current_price, current_high, current_low, current_time, 
                            times, closes, highs, lows)
            
            # 入场检查（需要间隔）
            if self.position_manager.can_open_position():
                diff = i - self.last_trade_bar
                if diff >= self.config.min_trade_interval:
                    self._check_entry(closes, i, current_time, current_open)
            
            # 更新最后交易K线（如果刚平仓）
        
        # 处理未平仓
        self._close_remaining(closes[-1], times[-1])
        
        return self.position_manager.trades, self.missed_signals
    
    def _reset(self):
        """重置状态"""
        self.position_manager = create_position_manager(self.config, self.risk_manager)
        self.last_trade_bar = -self.config.min_trade_interval
        self.missed_signals = []
    
    def _check_entry(self, closes: List[float], i: int, current_time: int, 
                     current_open: float):
        """检查入场 - 简化版"""
        # 简化版1-2-3形态检查（使用已传入的closes）
        pattern = self._find_123_pattern_simple(closes, i)
        if not pattern:
            return
        
        # 检查突破 - 提高阈值到0.5%
        breakout, thrust = self._check_breakout_simple(closes, pattern['p3'][0], 
                                                        'up' if pattern['type'] == 'low' else 'down')
        
        if not breakout or thrust < self.config.min_thrust:
            return
        
        # 生成信号
        direction = 'up' if pattern['type'] == 'low' else 'down'
        slippage = self.config.slippage_pct / 100
        
        if direction == 'up':
            entry_price = current_open * (1 + slippage)
            stop_loss = entry_price * (1 - self.config.stop_loss_pct / 100)
            take_profit = entry_price * (1 + self.config.take_profit_pct / 100)
            signal_type = 'long'
        else:
            entry_price = current_open * (1 - slippage)
            stop_loss = entry_price * (1 + self.config.stop_loss_pct / 100)
            take_profit = entry_price * (1 - self.config.take_profit_pct / 100)
            signal_type = 'short'
        
        # 检查相反方向持仓
        if self.position_manager.get_position_count() > 0:
            existing = self.position_manager.positions[0]
            if existing.type != signal_type:
                # 平仓相反方向
                self.position_manager.close_all_opposite_direction(
                    signal_type, current_open, slippage, current_time, i,
                    datetime.fromtimestamp(current_time/1000).strftime('%Y-%m-%d %H:%M')
                )
        
        # 开仓 - 创建简单signal对象
        from signals import PatternSignal
        signal = PatternSignal(
            signal_type=signal_type,
            pattern_name='1-2-3',
            entry_price=entry_price,
            stop_loss=stop_loss,
            take_profit=take_profit,
            thrust=thrust,
            confidence=0.7,
            metadata={}
        )
        
        pos = self.position_manager.open_position(
            signal, current_time, i, self.config.slippage_pct / 100
        )
        
        if pos:
            self.last_trade_bar = i
    
    def _check_exits(self, i: int, current_price: float, current_high: float,
                     current_low: float, current_time: int, times: List[int], 
                     closes: List[float], highs: List[float], lows: List[float]):
        """检查出场"""
        exited = []
        
        for pos in self.position_manager.positions:
            # 综合出场检查（传入ATR计算所需数据）
            exit_signal = self.risk_manager.check_exit(
                pos, current_price, current_high, current_low,
                highs=highs, lows=lows, closes=closes, idx=i
            )
            
            if exit_signal.should_exit:
                # 计算出场价（含滑点）
                slippage = self.config.slippage_pct / 100
                if pos.type == 'long':
                    exit_price = current_price * (1 - slippage)
                else:
                    exit_price = current_price * (1 + slippage)
                
                # 构建出场逻辑详情
                exit_logic = self._build_exit_logic(
                    pos, exit_signal, current_price, current_high, current_low
                )
                
                self.position_manager.close_position(
                    pos, exit_price, exit_signal.reason, exit_logic,
                    exit_time_str=datetime.fromtimestamp(current_time/1000).strftime('%Y-%m-%d %H:%M')
                )
                exited.append(pos)
        
        return exited
    
    def _build_exit_logic(self, pos, exit_signal, current_price, current_high, current_low) -> str:
        """构建出场逻辑详情"""
        reason = exit_signal.reason
        
        if "止损" in reason:
            if pos.type == 'long':
                return f"最低价{current_low:.2f}跌破止损价{pos.stop_loss:.2f}，触发止损"
            else:
                return f"最高价{current_high:.2f}涨破止损价{pos.stop_loss:.2f}，触发止损"
        elif "止盈" in reason:
            return f"收盘价{current_price:.2f}触及止盈价{pos.take_profit:.2f}，触发止盈"
        elif "超时" in reason:
            return f"持仓{pos.bars}根K线，超过最大持仓限制，强制平仓"
        elif "移动止损" in reason:
            return f"价格回撤，触发移动止损"
        elif "部分止盈" in reason:
            return f"盈利达到部分止盈点，止盈50%仓位"
        else:
            return reason
    
    def _close_remaining(self, last_price: float, last_time: int):
        """平掉剩余持仓"""
        slippage = self.config.slippage_pct / 100
        
        for pos in list(self.position_manager.positions):
            if pos.type == 'long':
                exit_price = last_price * (1 - slippage)
            else:
                exit_price = last_price * (1 + slippage)
            
            exit_logic = f"数据结束，未平仓，以收盘价{last_price:.2f}平仓"
            
            self.position_manager.close_position(
                pos, exit_price, '数据结束', exit_logic,
                exit_time_str=datetime.fromtimestamp(last_time/1000).strftime('%Y-%m-%d %H:%M')
            )
    
    def get_stats(self) -> Dict:
        """获取回测统计"""
        return self.position_manager.get_stats()
    
    def get_exit_reasons(self) -> Dict[str, int]:
        """获取出场原因统计"""
        return self.position_manager.get_exit_reasons()
    
    def reload_config(self, new_config: StrategyConfig):
        """
        重新加载配置（热更新）
        
        Args:
            new_config: 新的策略配置
        """
        new_config.validate()
        self.config = new_config
        
        # 重新初始化组件
        self.signal_generator = create_signal_generator(self.config)
        self.risk_manager = create_risk_manager(self.config)
        self.position_manager = create_position_manager(self.config, self.risk_manager)
    
    # ==================== 简化版信号函数 ====================
    
    def _find_123_pattern_simple(self, prices, current_idx):
        """简化版1-2-3形态"""
        max_lookback = 15
        if current_idx < 10 or current_idx - max_lookback < 0:
            return None
        
        for i in range(current_idx - 3, max(3, current_idx - max_lookback), -1):
            p1_price = prices[i]
            is_local_high = True
            is_local_low = True
            
            for j in range(max(0, i-3), i):
                if prices[j] >= p1_price:
                    is_local_high = False
                if prices[j] <= p1_price:
                    is_local_low = False
            
            if not is_local_high and not is_local_low:
                continue
            
            p2_idx = None
            for j in range(i+1, min(len(prices), i+5)):
                if is_local_high and prices[j] < prices[j-1]:
                    p2_idx = j
                    break
                if is_local_low and prices[j] > prices[j-1]:
                    p2_idx = j
                    break
            
            if not p2_idx:
                continue
            
            p3_idx = None
            for j in range(p2_idx+1, min(len(prices), p2_idx+5)):
                if is_local_high and prices[j] < prices[i]:
                    p3_idx = j
                    break
                if is_local_low and prices[j] > prices[i]:
                    p3_idx = j
                    break
            
            if p3_idx:
                return {'type': 'high' if is_local_high else 'low',
                        'p3': (p3_idx, prices[p3_idx])}
        
        return None
    
    def _check_breakout_simple(self, prices, hook_idx, direction):
        """简化版突破确认"""
        if hook_idx + 1 >= len(prices):
            return False, 0.0
        
        if direction == 'up' and prices[hook_idx + 1] > prices[hook_idx]:
            thrust = (prices[hook_idx + 1] - prices[hook_idx]) / prices[hook_idx] * 100
            return True, thrust
        elif direction == 'down' and prices[hook_idx + 1] < prices[hook_idx]:
            thrust = (prices[hook_idx] - prices[hook_idx + 1]) / prices[hook_idx] * 100
            return True, thrust
        return False, 0.0


def run_backtest(records: List[dict], config: StrategyConfig = None) -> tuple[List[Trade], List[dict]]:
    """
    快速回测函数
    """
    engine = BacktestEngine(config)
    return engine.run(records)


# ==================== Excel导出 ====================

def export_to_excel(trades: List[Trade], missed_signals: List[dict], filename: str,
                    config: StrategyConfig = None):
    """导出到Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    if config is None:
        config = StrategyConfig()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "交易记录"
    
    # 样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # 表头
    headers = ['序号', '方向', '入场时间', '出场时间', '入场价格', '出场价格', 
               '持仓K线', '同时持仓', '持仓金额', '计划止损', '计划止盈',
               '入场信号', '入场形态', '突破幅度%', '出场逻辑', '出场原因', 
               '手续费', '盈亏金额', '盈亏%', '余额']
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # 数据
    sl_pct = config.stop_loss_pct
    tp_pct = config.take_profit_pct
    
    for row, t in enumerate(trades, 2):
        ws.cell(row=row, column=1, value=row-1).border = thin_border
        ws.cell(row=row, column=2, value=t.position).border = thin_border
        ws.cell(row=row, column=3, value=t.entry_time_str).border = thin_border
        ws.cell(row=row, column=4, value=t.exit_time_str).border = thin_border
        ws.cell(row=row, column=5, value=t.entry_price).border = thin_border
        ws.cell(row=row, column=6, value=t.exit_price).border = thin_border
        ws.cell(row=row, column=7, value=t.hold_bars).border = thin_border
        ws.cell(row=row, column=8, value=t.concurrent_positions).border = thin_border
        ws.cell(row=row, column=9, value=round(t.position_size, 2)).border = thin_border
        
        # 计划止损止盈价格
        if t.entry_price > 0:
            sl_price = t.entry_price * (1 - sl_pct / 100)
            tp_price = t.entry_price * (1 + tp_pct / 100)
            ws.cell(row=row, column=10, value=round(sl_price, 6)).border = thin_border
            ws.cell(row=row, column=11, value=round(tp_price, 6)).border = thin_border
        
        ws.cell(row=row, column=12, value=t.entry_signal).border = thin_border
        ws.cell(row=row, column=13, value=t.entry_pattern).border = thin_border
        ws.cell(row=row, column=14, value=round(t.thrust, 2)).border = thin_border
        ws.cell(row=row, column=15, value=t.exit_logic).border = thin_border
        ws.cell(row=row, column=16, value=t.exit_reason).border = thin_border
        ws.cell(row=row, column=17, value=round(t.commission, 2)).border = thin_border
        
        # 盈亏颜色
        profit_cell = ws.cell(row=row, column=18, value=round(t.profit_usd, 2))
        profit_cell.border = thin_border
        if t.profit_usd > 0:
            profit_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif t.profit_usd < 0:
            profit_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        pnl_cell = ws.cell(row=row, column=19, value=round(t.pnl_pct, 2))
        pnl_cell.border = thin_border
        if t.pnl_pct > 0:
            pnl_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif t.pnl_pct < 0:
            pnl_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        ws.cell(row=row, column=20, value=round(t.balance_after, 2)).border = thin_border
    
    # 列宽
    for col in 'ABCDEFGHIJKLMNOPQRST':
        ws.column_dimensions[col].width = 12
    
    # 错过的信号 Sheet
    if missed_signals:
        ws2 = wb.create_sheet("错过的信号")
        
        header_fill2 = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        headers2 = ['序号', '时间', '入场信号', '入场形态', '突破幅度%', '价格', '未入场原因']
        
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill2
            cell.border = thin_border
        
        for row, sig in enumerate(missed_signals, 2):
            ws2.cell(row=row, column=1, value=row-1).border = thin_border
            ws2.cell(row=row, column=2, value=sig.get('time_str', '')).border = thin_border
            ws2.cell(row=row, column=3, value=sig.get('signal', '')).border = thin_border
            ws2.cell(row=row, column=4, value=sig.get('pattern', '')).border = thin_border
            ws2.cell(row=row, column=5, value=round(sig.get('thrust', 0), 2)).border = thin_border
            ws2.cell(row=row, column=6, value=sig.get('price', 0)).border = thin_border
            ws2.cell(row=row, column=7, value=sig.get('reason', '')).border = thin_border
    
    wb.save(filename)

    # ==================== 简化版信号函数 ====================
    
    def _find_123_pattern_simple(self, prices, current_idx):
        """简化版1-2-3形态"""
        max_lookback = 15
        if current_idx < 10 or current_idx - max_lookback < 0:
            return None
        
        for i in range(current_idx - 3, max(3, current_idx - max_lookback), -1):
            p1_price = prices[i]
            is_local_high = True
            is_local_low = True
            
            for j in range(max(0, i-3), i):
                if prices[j] >= p1_price:
                    is_local_high = False
                if prices[j] <= p1_price:
                    is_local_low = False
            
            if not is_local_high and not is_local_low:
                continue
            
            p2_idx = None
            for j in range(i+1, min(len(prices), i+5)):
                if is_local_high and prices[j] < prices[j-1]:
                    p2_idx = j
                    break
                if is_local_low and prices[j] > prices[j-1]:
                    p2_idx = j
                    break
            
            if not p2_idx:
                continue
            
            p3_idx = None
            for j in range(p2_idx+1, min(len(prices), p2_idx+5)):
                if is_local_high and prices[j] < prices[i]:
                    p3_idx = j
                    break
                if is_local_low and prices[j] > prices[i]:
                    p3_idx = j
                    break
            
            if p3_idx:
                return {'type': 'high' if is_local_high else 'low',
                        'p3': (p3_idx, prices[p3_idx])}
        
        return None
    
    def _check_breakout_simple(self, prices, hook_idx, direction):
        """简化版突破确认"""
        if hook_idx + 1 >= len(prices):
            return False, 0.0
        
        if direction == 'up' and prices[hook_idx + 1] > prices[hook_idx]:
            thrust = (prices[hook_idx + 1] - prices[hook_idx]) / prices[hook_idx] * 100
            return True, thrust
        elif direction == 'down' and prices[hook_idx + 1] < prices[hook_idx]:
            thrust = (prices[hook_idx] - prices[hook_idx + 1]) / prices[hook_idx] * 100
            return True, thrust
        return False, 0.0
