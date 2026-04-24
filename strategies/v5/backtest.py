"""
回测引擎模块 v5

v5 更新：
- 支持 skip_ross_hook 配置
- 支持 breakout_lookahead 参数
- 支持 max_daily_loss（每日最大亏损）
- 支持 KDJ/BOLL/EMA 过滤器
"""

from typing import List, Dict
from datetime import datetime

from config import StrategyConfig
from signals import create_signal_generator
from risk import create_risk_manager, ExitSignal
from position import create_position_manager, Trade, Position


class BacktestEngine:
    """回测引擎 - v5 增强版"""
    
    def __init__(self, config: StrategyConfig):
        """
        初始化回测引擎
        
        Args:
            config: 策略配置（由量化系统传入）
        """
        if config is None:
            raise ValueError("config 必须由量化系统传入，不能为None")
        
        self.config = config
        
        # 初始化组件
        self.signal_generator = create_signal_generator(self.config)
        self.risk_manager = create_risk_manager(self.config)
        self.position_manager = create_position_manager(self.config, self.risk_manager)
        
        # 状态
        self.last_trade_bar = -self.config.min_trade_interval
        self.missed_signals = []
        
        # 多周期数据
        self.ht_data = None  # 大周期数据
        
        # v5: 每日亏损追踪
        self.current_day = ""
    
    def load_higher_timeframe_data(self, ht_records: List[dict]):
        """
        加载大周期数据（用于多周期确认）
        
        Args:
            ht_records: 大周期K线数据列表
        """
        if ht_records:
            self.ht_data = {
                'closes': [r['close'] for r in ht_records],
                'highs': [r['high'] for r in ht_records],
                'lows': [r['low'] for r in ht_records],
            }
    
    def run(self, records: List[dict], ht_records: List[dict] = None) -> tuple[List[Trade], List[dict]]:
        """
        运行回测
        
        Args:
            records: K线数据列表
            ht_records: 大周期K线数据列表（可选，用于多周期确认）
        
        Returns:
            (交易列表, 错过的信号列表)
        """
        # 加载大周期数据
        if ht_records:
            self.load_higher_timeframe_data(ht_records)
        
        self._reset()
        
        # 提取数据
        closes = [r['close'] for r in records]
        opens = [r['open'] for r in records]
        highs = [r['high'] for r in records]
        lows = [r['low'] for r in records]
        times = [r['time'] for r in records]
        
        # 遍历K线
        for i in range(50, len(records)):
            current_price = closes[i]
            current_open = opens[i]
            current_high = highs[i]
            current_low = lows[i]
            current_time = times[i]
            
            # v5: 检查并重置每日亏损
            self._check_daily_reset(current_time)
            
            # v5: 检查每日最大亏损
            max_daily_loss = getattr(self.config, 'max_daily_loss', 50.0)
            if self.position_manager.daily_pnl <= -max_daily_loss:
                # 已达每日亏损上限，跳过入场
                continue
            
            # 更新持仓状态
            self.position_manager.update_positions(current_price)
            
            # 检查出场
            self._check_exits(i, current_price, current_high, current_low, current_time, 
                            times, closes, highs, lows)
            
            # 入场检查（需要间隔）
            # 注意：量化系统允许同向加仓，所以即使有持仓也可以开新仓
            # 但需要检查入场间隔
            diff = i - self.last_trade_bar
            if diff >= self.config.min_trade_interval:
                self._check_entry(records, closes, highs, lows, i, current_time, current_open)
        
        # 处理未平仓
        self._close_remaining(closes[-1], times[-1])
        
        return self.position_manager.trades, self.missed_signals
    
    def _reset(self):
        """重置状态"""
        self.position_manager = create_position_manager(self.config, self.risk_manager)
        self.last_trade_bar = -self.config.min_trade_interval
        self.missed_signals = []
        self.current_day = ""
        # 重置状态机
        self.signal_generator.reset_pattern_states()
    
    def _check_daily_reset(self, current_time: int):
        """v5: 检查并重置每日亏损"""
        current_day = datetime.fromtimestamp(current_time / 1000).strftime('%Y-%m-%d')
        if current_day != self.current_day:
            self.position_manager.reset_daily_pnl()
            self.current_day = current_day
    
    def _check_entry(self, records: List[dict], closes: List[float], 
                     highs: List[float], lows: List[float],
                     i: int, current_time: int, current_open: float):
        """检查入场 - v5 增强版（支持同向加仓）"""
        
        # 使用信号生成器
        ht_closes = self.ht_data['closes'] if self.ht_data else None
        
        # 预提取数据字典
        pre_extracted = {
            'closes': closes,
            'opens': [r['open'] for r in records],
            'highs': highs,
            'lows': lows,
            'volumes': [r.get('volume') for r in records],
        }
        
        signal = self.signal_generator.generate_signal(
            records, i, pre_extracted=pre_extracted
        )
        
        if not signal:
            return
        
        # 检查是否可以开仓（传入方向，允许同向加仓）
        if not self.position_manager.can_open_position(signal.side):
            return
        
        # 检查相反方向持仓，先平掉
        if self.position_manager.get_position_count() > 0:
            existing_positions = list(self.position_manager.positions)
            for existing in existing_positions:
                if existing.type != signal.side:
                    # 平仓相反方向
                    slippage = self.config.slippage_pct / 100
                    self.position_manager.close_all_opposite_direction(
                        signal.side, current_open, slippage, current_time, i,
                        datetime.fromtimestamp(current_time/1000).strftime('%Y-%m-%d %H:%M')
                    )
                    break  # 平掉一个反向持仓后继续
        
        # 开仓（允许同向加仓）
        pos = self.position_manager.open_position(
            signal, current_time, i
        )
        
        if pos:
            self.last_trade_bar = i
    
    def _check_exits(self, i: int, current_price: float, current_high: float,
                     current_low: float, current_time: int, times: List[int], 
                     closes: List[float], highs: List[float], lows: List[float]):
        """检查出场"""
        exited = []
        
        for pos in self.position_manager.positions:
            # 综合出场检查（传入ATR计算所需数据）
            exit_signal = self.risk_manager.check_exit(
                pos, current_price, current_high, current_low,
                highs=highs, lows=lows, closes=closes, idx=i
            )
            
            if exit_signal.should_exit:
                # 计算出场价（含滑点）
                slippage = self.config.slippage_pct / 100
                if pos.type == 'long':
                    exit_price = current_price * (1 - slippage)
                else:
                    exit_price = current_price * (1 + slippage)
                
                # 构建出场逻辑详情
                exit_logic = self._build_exit_logic(
                    pos, exit_signal, current_price, current_high, current_low
                )
                
                self.position_manager.close_position(
                    pos, exit_price, exit_signal.reason, exit_logic,
                    exit_time_str=datetime.fromtimestamp(current_time/1000).strftime('%Y-%m-%d %H:%M')
                )
                exited.append(pos)
        
        return exited
    
    def _build_exit_logic(self, pos: Position, exit_signal: ExitSignal, 
                          current_price: float, current_high: float, 
                          current_low: float) -> str:
        """构建出场逻辑详情"""
        reason = exit_signal.reason
        
        if "止损" in reason:
            if pos.type == 'long':
                return f"最低价{current_low:.2f}跌破止损价{pos.stop_loss:.2f}，触发止损"
            else:
                return f"最高价{current_high:.2f}涨破止损价{pos.stop_loss:.2f}，触发止损"
        elif "止盈" in reason:
            return f"收盘价{current_price:.2f}触及止盈价{pos.take_profit:.2f}，触发止盈"
        elif "超时" in reason:
            return f"持仓{pos.bars}根K线，超过最大持仓限制，强制平仓"
        elif "移动止损" in reason:
            return f"价格回撤，触发移动止损"
        elif "部分止盈" in reason:
            return f"盈利达到部分止盈点，止盈50%仓位"
        elif "ATR止损" in reason:
            return f"价格跌破ATR止损线，触发ATR止损"
        else:
            return reason
    
    def _close_remaining(self, last_price: float, last_time: int):
        """平掉剩余持仓"""
        slippage = self.config.slippage_pct / 100
        
        for pos in list(self.position_manager.positions):
            if pos.type == 'long':
                exit_price = last_price * (1 - slippage)
            else:
                exit_price = last_price * (1 + slippage)
            
            exit_logic = f"数据结束，未平仓，以收盘价{last_price:.2f}平仓"
            
            self.position_manager.close_position(
                pos, exit_price, '数据结束', exit_logic,
                exit_time_str=datetime.fromtimestamp(last_time/1000).strftime('%Y-%m-%d %H:%M')
            )
    
    def get_stats(self) -> Dict:
        """获取回测统计"""
        return self.position_manager.get_stats()
    
    def get_exit_reasons(self) -> Dict[str, int]:
        """获取出场原因统计"""
        return self.position_manager.get_exit_reasons()
    
    def reload_config(self, new_config: StrategyConfig):
        """
        重新加载配置（热更新）
        
        Args:
            new_config: 新的策略配置
        """
        new_config.validate()
        self.config = new_config
        
        # 重新初始化组件
        self.signal_generator = create_signal_generator(self.config)
        self.risk_manager = create_risk_manager(self.config)
        self.position_manager = create_position_manager(self.config, self.risk_manager)


def run_backtest(records: List[dict], config: StrategyConfig = None) -> tuple[List[Trade], List[dict]]:
    """
    快速回测函数
    """
    engine = BacktestEngine(config)
    return engine.run(records)


# ==================== Excel导出 ====================

def export_to_excel(trades: List[Trade], missed_signals: List[dict], filename: str,
                    config: StrategyConfig = None):
    """导出到Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    
    if config is None:
        config = StrategyConfig()
    
    # 按入场时间排序
    sorted_trades = sorted(trades, key=lambda t: t.entry_time_str)
    
    # 重新计算余额（排序后）
    initial_balance = config.initial_balance
    running_balance = initial_balance
    for t in sorted_trades:
        running_balance += t.profit_usd
        t.balance_after = running_balance
    
    # 计算累计统计
    wins = sum(1 for t in sorted_trades if t.profit_usd > 0)
    losses = sum(1 for t in sorted_trades if t.profit_usd < 0)
    total_trades = len(sorted_trades)
    win_rate = wins / total_trades * 100 if total_trades > 0 else 0
    total_pnl = sum(t.profit_usd for t in sorted_trades)
    final_balance = initial_balance + total_pnl
    
    wb = Workbook()
    ws = wb.active
    ws.title = "交易记录"
    
    # 样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # 表头 - 按用户要求的顺序
    headers = ['序号', '入场形态', '入场时间', '入场价格', '方向', '持仓金额', '盈亏金额',
               '持仓K线', '出场时间', '出场价格', '突破幅度%', '计划止损', '计划止盈',
               '出场原因', '手续费', '余额']
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # 数据 - 按新字段顺序填充
    sl_pct = config.stop_loss_pct
    tp_pct = config.take_profit_pct
    
    for row, t in enumerate(sorted_trades, 2):
        # 1. 序号
        ws.cell(row=row, column=1, value=row-1).border = thin_border
        # 2. 入场形态
        ws.cell(row=row, column=2, value=t.entry_pattern).border = thin_border
        # 3. 入场时间
        ws.cell(row=row, column=3, value=t.entry_time_str).border = thin_border
        # 4. 入场价格
        ws.cell(row=row, column=4, value=t.entry_price).border = thin_border
        # 5. 方向
        ws.cell(row=row, column=5, value=t.position).border = thin_border
        # 6. 持仓金额
        ws.cell(row=row, column=6, value=round(t.position_size, 2)).border = thin_border
        # 7. 盈亏金额（带颜色）
        profit_cell = ws.cell(row=row, column=7, value=round(t.profit_usd, 2))
        profit_cell.border = thin_border
        if t.profit_usd > 0:
            profit_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif t.profit_usd < 0:
            profit_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        # 8. 持仓K线
        ws.cell(row=row, column=8, value=t.hold_bars).border = thin_border
        # 9. 出场时间
        ws.cell(row=row, column=9, value=t.exit_time_str).border = thin_border
        # 10. 出场价格
        ws.cell(row=row, column=10, value=t.exit_price).border = thin_border
        # 11. 突破幅度%
        ws.cell(row=row, column=11, value=round(t.thrust, 2)).border = thin_border
        # 12. 计划止损（价格）
        if t.entry_price > 0:
            sl_price = t.entry_price * (1 - sl_pct / 100)
            ws.cell(row=row, column=12, value=round(sl_price, 6)).border = thin_border
        else:
            ws.cell(row=row, column=12, value='').border = thin_border
        # 13. 计划止盈（价格）
        if t.entry_price > 0:
            tp_price = t.entry_price * (1 + tp_pct / 100)
            ws.cell(row=row, column=13, value=round(tp_price, 6)).border = thin_border
        else:
            ws.cell(row=row, column=13, value='').border = thin_border
        # 14. 出场原因
        ws.cell(row=row, column=14, value=t.exit_reason).border = thin_border
        # 15. 手续费
        ws.cell(row=row, column=15, value=round(t.commission, 2)).border = thin_border
        # 16. 余额
        ws.cell(row=row, column=16, value=round(t.balance_after, 2)).border = thin_border
    
    # 列宽
    for col in 'ABCDEFGHIJKLMNOPQRST':
        ws.column_dimensions[col].width = 12
    
    # 错过的信号 Sheet
    if missed_signals:
        ws2 = wb.create_sheet("错过的信号")
        
        header_fill2 = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        headers2 = ['序号', '时间', '入场信号', '入场形态', '突破幅度%', '价格', '未入场原因']
        
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill2
            cell.border = thin_border
        
        for row, m in enumerate(missed_signals, 2):
            ws2.cell(row=row, column=1, value=row-1).border = thin_border
            ws2.cell(row=row, column=2, value=m.get('time', '')).border = thin_border
            ws2.cell(row=row, column=3, value=m.get('side', '')).border = thin_border
            ws2.cell(row=row, column=4, value=m.get('pattern_name', '')).border = thin_border
            ws2.cell(row=row, column=5, value=round(m.get('thrust', 0), 2)).border = thin_border
            ws2.cell(row=row, column=6, value=m.get('price', 0)).border = thin_border
            ws2.cell(row=row, column=7, value=m.get('reason', '')).border = thin_border
    
    # 配置参数 Sheet
    ws3 = wb.create_sheet("配置参数")
    config_dict = config.to_dict() if hasattr(config, 'to_dict') else {}
    
    header_fill3 = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    headers3 = ['参数名', '参数值']
    
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill3
        cell.border = thin_border
    
    for row, (key, value) in enumerate(config_dict.items(), 2):
        ws3.cell(row=row, column=1, value=key).border = thin_border
        ws3.cell(row=row, column=2, value=value).border = thin_border
    
    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 15
    
    # 统计摘要 Sheet
    ws4 = wb.create_sheet("统计摘要")
    header_fill4 = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
    headers4 = ['指标', '数值']
    for col, h in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill4
        cell.border = thin_border
    
    stats_data = [
        ('总交易数', total_trades),
        ('盈利交易', wins),
        ('亏损交易', losses),
        ('胜率', f'{win_rate:.1f}%'),
        ('总盈亏', f'{total_pnl:.2f} USDT'),
        ('最终余额', f'{final_balance:.2f} USDT'),
        ('初始余额', f'{initial_balance:.2f} USDT'),
        ('收益率', f'{total_pnl/initial_balance*100:.1f}%'),
    ]
    for row, (key, value) in enumerate(stats_data, 2):
        ws4.cell(row=row, column=1, value=key).border = thin_border
        ws4.cell(row=row, column=2, value=value).border = thin_border
    ws4.column_dimensions['A'].width = 15
    ws4.column_dimensions['B'].width = 18
    
    # 保存
    wb.save(filename)
    print(f"✅ Excel已保存: {filename}")