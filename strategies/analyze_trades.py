#!/usr/bin/env python3
"""
BTC交易分析脚本
基于RSI、KDJ、BOLL指标分析交易机会
"""

import pymongo
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

# MongoDB连接
client = pymongo.MongoClient('mongodb://localhost:27017/')
db = client['trading-data']
collection = db['btc_5m']

# 获取所有数据，按时间正序排列
records = list(collection.find().sort('time', 1).limit(1000))
for r in records:
    r.pop('_id', None)

print(f"获取到 {len(records)} 条数据")
print(f"时间范围: {datetime.fromtimestamp(records[0]['time']/1000)} 到 {datetime.fromtimestamp(records[-1]['time']/1000)}")

# 交易参数
TRADE_AMOUNT = 100  # 每次开仓100美元
TRADES = []

def calculate_profit(entry_price, exit_price, is_long=True):
    """计算盈亏"""
    if is_long:
        return (exit_price - entry_price) / entry_price * TRADE_AMOUNT
    else:
        return (entry_price - exit_price) / entry_price * TRADE_AMOUNT

# 预计算前一根K线的指标，用于判断交叉
prev_rsi = None
prev_k = None
prev_d = None
position = None  # None, 'long', 'short'
entry_price = 0

for i in range(1, len(records)):
    curr = records[i]
    prev = records[i-1]
    
    curr_rsi = curr.get('rsi')
    curr_k = curr.get('kdj', {}).get('k')
    curr_d = curr.get('kdj', {}).get('d')
    curr_boll = curr.get('boll', {})
    
    prev_rsi_val = prev.get('rsi')
    prev_k_val = prev.get('kdj', {}).get('k')
    prev_d_val = prev.get('kdj', {}).get('d')
    prev_boll = prev.get('boll', {})
    
    # 跳过无效数据
    if not all([curr_rsi, curr_k, curr_d, curr_boll.get('upper')]):
        continue
    
    trade_trigger = None
    trade_reason = ""
    
    # === 买入信号 ===
    # 1. RSI超卖回升: RSI从<30回升到>30
    if prev_rsi_val and prev_rsi_val < 30 and curr_rsi > 30 and position is None:
        trade_trigger = "RSI超卖回升"
        trade_reason = f"RSI: {prev_rsi_val:.2f} -> {curr_rsi:.2f}"
        position = 'long'
        entry_price = curr['price']
        
    # 2. KDJ金叉: K线从下往上穿过D线
    elif prev_k_val and prev_d_val and prev_k_val < prev_d_val and curr_k > curr_d and position is None:
        trade_trigger = "KDJ金叉"
        trade_reason = f"K: {prev_k_val:.2f}->{curr_k:.2f}, D: {prev_d_val:.2f}->{curr_d:.2f}"
        position = 'long'
        entry_price = curr['price']
        
    # 3. BOLL下轨支撑: 价格从下轨下方回到轨内
    elif prev_boll.get('lower') and curr['price'] > curr_boll['lower'] and prev['price'] < prev_boll['lower'] and position is None:
        trade_trigger = "BOLL下轨支撑"
        trade_reason = f"价格突破下轨 {prev_boll['lower']:.2f}"
        position = 'long'
        entry_price = curr['price']
    
    # === 卖出信号 ===
    # 1. RSI超买回落: RSI从>70回落
    elif prev_rsi_val and prev_rsi_val > 70 and curr_rsi < 70 and position == 'long':
        trade_trigger = "RSI超买回落"
        trade_reason = f"RSI: {prev_rsi_val:.2f} -> {curr_rsi:.2f}"
        profit = calculate_profit(entry_price, curr['price'], is_long=True)
        TRADES.append({
            'time': datetime.fromtimestamp(curr['time']/1000).strftime('%Y-%m-%d %H:%M:%S'),
            'type': '做多',
            'entry_price': entry_price,
            'exit_price': curr['price'],
            'trigger': trade_trigger,
            'reason': trade_reason,
            'profit': profit,
            'profit_usd': f"${profit:.2f}"
        })
        position = None
        
    # 2. KDJ死叉: K线从上往下穿过D线
    elif prev_k_val and prev_d_val and prev_k_val > prev_d_val and curr_k < curr_d and position == 'long':
        trade_trigger = "KDJ死叉"
        trade_reason = f"K: {prev_k_val:.2f}->{curr_k:.2f}, D: {prev_d_val:.2f}->{curr_d:.2f}"
        profit = calculate_profit(entry_price, curr['price'], is_long=True)
        TRADES.append({
            'time': datetime.fromtimestamp(curr['time']/1000).strftime('%Y-%m-%d %H:%M:%S'),
            'type': '做多',
            'entry_price': entry_price,
            'exit_price': curr['price'],
            'trigger': trade_trigger,
            'reason': trade_reason,
            'profit': profit,
            'profit_usd': f"${profit:.2f}"
        })
        position = None
        
    # 3. BOLL上轨压力: 价格从上轨上方回到轨内
    elif prev_boll.get('upper') and curr['price'] < curr_boll['upper'] and prev['price'] > prev_boll['upper'] and position == 'long':
        trade_trigger = "BOLL上轨压力"
        trade_reason = f"价格跌破上轨 {prev_boll['upper']:.2f}"
        profit = calculate_profit(entry_price, curr['price'], is_long=True)
        TRADES.append({
            'time': datetime.fromtimestamp(curr['time']/1000).strftime('%Y-%m-%d %H:%M:%S'),
            'type': '做多',
            'entry_price': entry_price,
            'exit_price': curr['price'],
            'trigger': trade_trigger,
            'reason': trade_reason,
            'profit': profit,
            'profit_usd': f"${profit:.2f}"
        })
        position = None

# 如果还有持仓，平仓
if position == 'long':
    last_price = records[-1]['price']
    profit = calculate_profit(entry_price, last_price, is_long=True)
    TRADES.append({
        'time': datetime.fromtimestamp(records[-1]['time']/1000).strftime('%Y-%m-%d %H:%M:%S'),
        'type': '做多(持仓平仓)',
        'entry_price': entry_price,
        'exit_price': last_price,
        'trigger': '数据结束',
        'reason': '自动平仓',
        'profit': profit,
        'profit_usd': f"${profit:.2f}"
    })

print(f"\n共发现 {len(TRADES)} 笔交易")

# 写入Excel
wb = Workbook()
ws = wb.active
ws.title = "BTC交易记录"

# 表头样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

headers = ['时间', '交易类型', '入场价格', '出场价格', '触发条件', '技术理由', '盈亏(%)', '盈亏($)']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# 写入数据
profit_total = 0
win_count = 0

for row, trade in enumerate(TRADES, 2):
    ws.cell(row=row, column=1, value=trade['time'])
    ws.cell(row=row, column=2, value=trade['type'])
    ws.cell(row=row, column=3, value=trade['entry_price'])
    ws.cell(row=row, column=4, value=trade['exit_price'])
    ws.cell(row=row, column=5, value=trade['trigger'])
    ws.cell(row=row, column=6, value=trade['reason'])
    
    profit_pct = (trade['exit_price'] - trade['entry_price']) / trade['entry_price'] * 100
    ws.cell(row=row, column=7, value=f"{profit_pct:.2f}%")
    ws.cell(row=row, column=8, value=trade['profit_usd'])
    
    profit_total += trade['profit']
    if trade['profit'] > 0:
        win_count += 1
    
    # 盈利绿色，亏损红色
    profit_cell = ws.cell(row=row, column=8)
    if trade['profit'] >= 0:
        profit_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        profit_cell.font = Font(color="006100")
    else:
        profit_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        profit_cell.font = Font(color="9C0006")

# 统计行
ws.cell(row=len(TRADES)+2, column=1, value="总计")
ws.cell(row=len(TRADES)+2, column=7, value=f"总盈亏: ${profit_total:.2f}")
ws.cell(row=len(TRADES)+2, column=8, value=f"胜率: {win_count}/{len(TRADES)} = {win_count/len(TRADES)*100:.1f}%")

# 调整列宽
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 35
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 12

# 保存文件
output_path = '/Users/makaihong/.openclaw/workspaces/ross/BTC_交易记录_5m.xlsx'
wb.save(output_path)
print(f"\n交易记录已保存到: {output_path}")
print(f"总盈亏: ${profit_total:.2f}")
print(f"胜率: {win_count}/{len(TRADES)} = {win_count/len(TRADES)*100:.1f}%")
