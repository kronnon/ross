#!/usr/bin/env python3
"""
洛氏霍克交易系统 v1.0
核心策略：多指标共振 + 动态止盈止损
"""

import pymongo
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json

# ==================== 策略参数配置 ====================
class StrategyConfig:
    def __init__(self):
        # 基础参数
        self.trade_amount = 100  # 每次开仓金额(USD)
        
        # RSI参数
        self.rsi_oversold = 30    # RSI超卖阈值
        self.rsi_overbought = 70  # RSI超买阈值
        
        # KDJ参数
        self.kdj_oversold = 20    # KDJ超卖阈值
        self.kdj_overbought = 80 # KDJ超买阈值
        
        # BOLL参数
        self.boll_period = 20     # BOLL周期
        self.boll_std = 2         # BOLL标准差倍数
        
        # 止盈止损
        self.stop_loss_pct = 0.5  # 止损百分比
        self.take_profit_pct = 1.0 # 止盈百分比
        
        # 持仓时间限制
        self.max_hold_bars = 24  # 最大持仓K线数（5m*24=2小时）
        
        # 最小间隔（避免频繁交易）
        self.min_trade_interval = 3  # 最小交易间隔(K线数)
        
    def to_dict(self):
        return {k: v for k, v in self.__dict__.items() if not k.startswith('_')}


class RockyHedgehogTradingSystem:
    def __init__(self, config=None):
        self.config = config or StrategyConfig()
        self.trades = []
        self.position = None  # None, 'long'
        self.entry_price = 0
        self.entry_time = 0
        self.bars_since_entry = 0
        self.last_trade_bar = -999
        
    def reset(self):
        """重置交易状态"""
        self.trades = []
        self.position = None
        self.entry_price = 0
        self.entry_time = 0
        self.bars_since_entry = 0
        self.last_trade_bar = -999
        
    def check_entry_signal(self, curr, prev, bar_index):
        """检查入场信号"""
        # 检查最小间隔
        if bar_index - self.last_trade_bar < self.config.min_trade_interval:
            return None, ""
            
        rsi = curr.get('rsi')
        k = curr.get('kdj', {}).get('k')
        d = curr.get('kdj', {}).get('d')
        j = curr.get('kdj', {}).get('j')
        boll = curr.get('boll', {})
        
        prev_rsi = prev.get('rsi')
        prev_k = prev.get('kdj', {}).get('k')
        prev_d = prev.get('kdj', {}).get('d')
        prev_boll = prev.get('boll', {})
        
        if not all([rsi, k, d, boll.get('upper')]):
            return None, ""
            
        signals = []
        
        # 1. RSI超卖回升
        if prev_rsi and prev_rsi < self.config.rsi_oversold and rsi > self.config.rsi_oversold:
            signals.append("RSI超卖")
            
        # 2. KDJ金叉
        if prev_k and prev_d and prev_k < prev_d and k > d:
            signals.append("KDJ金叉")
            
        # 3. BOLL下轨支撑
        if prev_boll.get('lower') and prev['price'] < prev_boll['lower'] and curr['price'] > boll['lower']:
            signals.append("BOLL支撑")
            
        # 4. KDJ超卖
        if k < self.config.kdj_oversold:
            signals.append("KDJ超卖")
            
        # 多指标共振：至少2个信号
        if len(signals) >= 2:
            return 'long', " + ".join(signals)
            
        return None, ""
        
    def check_exit_signal(self, curr, prev, bar_index):
        """检查出场信号"""
        if self.position != 'long':
            return None, ""
            
        rsi = curr.get('rsi')
        k = curr.get('kdj', {}).get('k')
        d = curr.get('kdj', {}).get('d')
        boll = curr.get('boll', {})
        
        prev_rsi = prev.get('rsi')
        prev_k = prev.get('kdj', {}).get('k')
        prev_d = prev.get('kdj', {}).get('d')
        prev_boll = prev.get('boll', {})
        
        if not all([rsi, k, d, boll.get('upper')]):
            return None, ""
            
        exit_reason = ""
        
        # 1. 止损
        pnl_pct = (curr['price'] - self.entry_price) / self.entry_price * 100
        if pnl_pct <= -self.config.stop_loss_pct:
            return 'stop_loss', f"止损 {pnl_pct:.2f}%"
            
        # 2. 止盈
        if pnl_pct >= self.config.take_profit_pct:
            return 'take_profit', f"止盈 {pnl_pct:.2f}%"
            
        # 3. KDJ死叉
        if prev_k and prev_d and prev_k > prev_d and k < d:
            return 'kdj_death', "KDJ死叉"
            
        # 4. RSI超买回落
        if prev_rsi and prev_rsi > self.config.rsi_overbought and rsi < self.config.rsi_overbought:
            return 'rsi_overbought', "RSI超买"
            
        # 5. BOLL上轨压力
        if prev_boll.get('upper') and prev['price'] > prev_boll['upper'] and curr['price'] < boll['upper']:
            return 'boll_pressure', "BOLL压力"
            
        # 6. 持仓时间到限
        self.bars_since_entry += 1
        if self.bars_since_entry >= self.config.max_hold_bars:
            return 'time_up', f"持仓超时({self.bars_since_entry}根)"
            
        return None, ""
        
    def run_backtest(self, records, verbose=True):
        """运行回测"""
        self.reset()
        
        for i in range(1, len(records)):
            curr = records[i]
            prev = records[i-1]
            
            # 入场检查
            if self.position is None:
                signal, reason = self.check_entry_signal(curr, prev, i)
                if signal:
                    self.position = 'long'
                    self.entry_price = curr['price']
                    self.entry_time = curr['time']
                    self.bars_since_entry = 0
                    self.last_trade_bar = i
                    
            # 出场检查
            else:
                exit_signal, exit_reason = self.check_exit_signal(curr, prev, i)
                if exit_signal:
                    profit = (curr['price'] - self.entry_price) / self.entry_price * self.config.trade_amount
                    pnl_pct = (curr['price'] - self.entry_price) / self.entry_price * 100
                    
                    self.trades.append({
                        'entry_time': datetime.fromtimestamp(self.entry_time/1000).strftime('%Y-%m-%d %H:%M'),
                        'exit_time': datetime.fromtimestamp(curr['time']/1000).strftime('%Y-%m-%d %H:%M'),
                        'entry_price': self.entry_price,
                        'exit_price': curr['price'],
                        'hold_bars': self.bars_since_entry,
                        'exit_reason': exit_reason,
                        'profit_usd': profit,
                        'pnl_pct': pnl_pct
                    })
                    self.position = None
                    
        # 最后持仓平仓
        if self.position == 'long':
            last = records[-1]
            profit = (last['price'] - self.entry_price) / self.entry_price * self.config.trade_amount
            pnl_pct = (last['price'] - self.entry_price) / self.entry_price * 100
            self.trades.append({
                'entry_time': datetime.fromtimestamp(self.entry_time/1000).strftime('%Y-%m-%d %H:%M'),
                'exit_time': datetime.fromtimestamp(last['time']/1000).strftime('%Y-%m-%d %H:%M'),
                'entry_price': self.entry_price,
                'exit_price': last['price'],
                'hold_bars': self.bars_since_entry,
                'exit_reason': '数据结束',
                'profit_usd': profit,
                'pnl_pct': pnl_pct
            })
            
        return self.get_stats()
        
    def get_stats(self):
        """获取统计数据"""
        if not self.trades:
            return {
                'trade_count': 0,
                'win_rate': 0,
                'avg_profit': 0,
                'avg_loss': 0,
                'profit_factor': 0,
                'total_pnl': 0,
                'max_consecutive_wins': 0,
                'max_consecutive_losses': 0
            }
            
        wins = [t for t in self.trades if t['profit_usd'] > 0]
        losses = [t for t in self.trades if t['profit_usd'] <= 0]
        
        win_rate = len(wins) / len(self.trades) * 100
        
        avg_profit = sum(t['profit_usd'] for t in wins) / len(wins) if wins else 0
        avg_loss = sum(t['profit_usd'] for t in losses) / len(losses) if losses else 0
        
        total_win = sum(t['profit_usd'] for t in wins)
        total_loss = abs(sum(t['profit_usd'] for t in losses))
        profit_factor = total_win / total_loss if total_loss > 0 else float('inf')
        
        # 连胜连负
        max_win_streak = 0
        max_loss_streak = 0
        current_win_streak = 0
        current_loss_streak = 0
        
        for t in self.trades:
            if t['profit_usd'] > 0:
                current_win_streak += 1
                current_loss_streak = 0
                max_win_streak = max(max_win_streak, current_win_streak)
            else:
                current_loss_streak += 1
                current_win_streak = 0
                max_loss_streak = max(max_loss_streak, current_loss_streak)
                
        return {
            'trade_count': len(self.trades),
            'win_rate': win_rate,
            'avg_profit': avg_profit,
            'avg_loss': avg_loss,
            'profit_factor': profit_factor,
            'total_pnl': sum(t['profit_usd'] for t in self.trades),
            'max_consecutive_wins': max_win_streak,
            'max_consecutive_losses': max_loss_streak
        }


def load_data(limit=2000):
    """加载数据"""
    client = pymongo.MongoClient('mongodb://localhost:27017/')
    db = client['trading-data']
    collection = db['btc_5m']
    records = list(collection.find().sort('time', 1).limit(limit))
    for r in records:
        r.pop('_id', None)
    return records


def optimize_parameters(records):
    """参数优化"""
    print("\n" + "="*60)
    print("开始参数优化...")
    print("="*60)
    
    best_score = -float('inf')
    best_config = None
    best_stats = None
    
    # 参数网格
    rsi_oversold_range = [25, 30, 35]
    rsi_overbought_range = [65, 70, 75]
    stop_loss_range = [0.3, 0.5, 0.7]
    take_profit_range = [0.8, 1.0, 1.2, 1.5]
    max_hold_range = [12, 24, 36]
    min_interval_range = [2, 3, 5]
    
    results = []
    
    for rsi_oversold in rsi_oversold_range:
        for rsi_overbought in rsi_overbought_range:
            if rsi_oversold >= rsi_overbought:
                continue
            for stop_loss in stop_loss_range:
                for take_profit in take_profit_range:
                    if take_profit <= stop_loss:
                        continue
                    for max_hold in max_hold_range:
                        for min_interval in min_interval_range:
                            config = StrategyConfig()
                            config.rsi_oversold = rsi_oversold
                            config.rsi_overbought = rsi_overbought
                            config.stop_loss_pct = stop_loss
                            config.take_profit_pct = take_profit
                            config.max_hold_bars = max_hold
                            config.min_trade_interval = min_interval
                            
                            system = RockyHedgehogTradingSystem(config)
                            stats = system.run_backtest(records, verbose=False)
                            
                            if stats['trade_count'] == 0:
                                continue
                                
                            # 综合评分：考虑胜率、盈亏比、交易频率
                            # 目标：较高的胜率、较好的盈亏比、适中的交易频率
                            win_rate_score = stats['win_rate'] / 100 * 40  # 胜率权重40
                            
                            # 盈亏比评分
                            profit_factor = stats['profit_factor'] if stats['profit_factor'] != float('inf') else 5
                            pnl_score = min(profit_factor / 3, 1) * 30  # 盈亏比权重30
                            
                            # 交易频率评分：每天1-3笔为最佳
                            days = 3.5  # 约3.5天数据
                            trades_per_day = stats['trade_count'] / days
                            if 1 <= trades_per_day <= 3:
                                freq_score = 30
                            elif trades_per_day < 1:
                                freq_score = trades_per_day / 1 * 20
                            else:
                                freq_score = max(0, 30 - (trades_per_day - 3) * 5)
                            
                            total_score = win_rate_score + pnl_score + freq_score
                            
                            results.append({
                                'config': config,
                                'stats': stats,
                                'score': total_score
                            })
                            
                            if total_score > best_score:
                                best_score = total_score
                                best_config = config
                                best_stats = stats
    
    # 排序结果
    results.sort(key=lambda x: x['score'], reverse=True)
    
    print("\n=== 优化结果 TOP 5 ===")
    for i, r in enumerate(results[:5], 1):
        print(f"\n#{i} 评分: {r['score']:.1f}")
        print(f"   交易次数: {r['stats']['trade_count']}, 胜率: {r['stats']['win_rate']:.1f}%")
        print(f"   盈亏比: {r['stats']['profit_factor']:.2f}, 总盈亏: ${r['stats']['total_pnl']:.2f}")
        print(f"   参数: RSI({r['config'].rsi_oversold}/{r['config'].rsi_overbought}), "
              f"止损{r['config'].stop_loss_pct}%, 止盈{r['config'].take_profit_pct}%, "
              f"持仓{r['config'].max_hold_bars}根")
    
    return best_config, best_stats, results[:10]


def run_final_test(records, config):
    """运行最终测试并生成Excel"""
    print("\n" + "="*60)
    print("运行最终测试...")
    print("="*60)
    
    system = RockyHedgehogTradingSystem(config)
    stats = system.run_backtest(records)
    
    print(f"\n最终统计:")
    print(f"  交易次数: {stats['trade_count']}")
    print(f"  胜率: {stats['win_rate']:.1f}%")
    print(f"  平均盈利: ${stats['avg_profit']:.2f}")
    print(f"  平均亏损: ${stats['avg_loss']:.2f}")
    print(f"  盈亏比: {stats['profit_factor']:.2f}")
    print(f"  总盈亏: ${stats['total_pnl']:.2f}")
    print(f"  最大连胜: {stats['max_consecutive_wins']}")
    print(f"  最大连亏: {stats['max_consecutive_losses']}")
    
    # 生成Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "交易记录"
    
    # 表头
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    headers = ['入场时间', '出场时间', '入场价', '出场价', '持仓(K线)', '出场理由', '盈亏($)', '盈亏(%)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
    
    # 数据
    for row, trade in enumerate(system.trades, 2):
        ws.cell(row=row, column=1, value=trade['entry_time'])
        ws.cell(row=row, column=2, value=trade['exit_time'])
        ws.cell(row=row, column=3, value=trade['entry_price'])
        ws.cell(row=row, column=4, value=trade['exit_price'])
        ws.cell(row=row, column=5, value=trade['hold_bars'])
        ws.cell(row=row, column=6, value=trade['exit_reason'])
        ws.cell(row=row, column=7, value=round(trade['profit_usd'], 2))
        ws.cell(row=row, column=8, value=f"{trade['pnl_pct']:.2f}%")
        
        # 颜色
        profit_cell = ws.cell(row=row, column=7)
        if trade['profit_usd'] >= 0:
            profit_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            profit_cell.font = Font(color="006100")
        else:
            profit_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            profit_cell.font = Font(color="9C0006")
    
    # 统计
    stats_row = len(system.trades) + 3
    ws.cell(row=stats_row, column=1, value="总计").font = Font(bold=True)
    ws.cell(row=stats_row, column=7, value=f"${stats['total_pnl']:.2f}").font = Font(bold=True)
    ws.cell(row=stats_row+1, column=1, value=f"交易次数").font = Font(bold=True)
    ws.cell(row=stats_row+1, column=7, value=stats['trade_count'])
    ws.cell(row=stats_row+2, column=1, value=f"胜率").font = Font(bold=True)
    ws.cell(row=stats_row+2, column=7, value=f"{stats['win_rate']:.1f}%")
    ws.cell(row=stats_row+3, column=1, value=f"盈亏比").font = Font(bold=True)
    ws.cell(row=stats_row+3, column=7, value=f"{stats['profit_factor']:.2f}")
    
    # 列宽
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 16
    
    output = '/Users/makaihong/.openclaw/workspaces/ross/洛氏霍克交易系统_v1.xlsx'
    wb.save(output)
    print(f"\n结果已保存: {output}")
    
    return stats


if __name__ == "__main__":
    # 加载数据
    records = load_data(2000)
    print(f"加载数据: {len(records)} 条")
    print(f"时间范围: {datetime.fromtimestamp(records[0]['time']/1000)} ~ {datetime.fromtimestamp(records[-1]['time']/1000)}")
    
    # 优化参数
    best_config, best_stats, top_results = optimize_parameters(records)
    
    # 运行最终测试
    stats = run_final_test(records, best_config)
    
    # 保存配置
    config_file = '/Users/makaihong/.openclaw/workspaces/ross/trading_config.json'
    with open(config_file, 'w') as f:
        json.dump({
            'config': best_config.to_dict(),
            'stats': best_stats
        }, f, indent=2)
    print(f"\n配置已保存: {config_file}")
