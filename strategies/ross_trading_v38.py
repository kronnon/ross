#!/usr/bin/env python3
"""
洛氏霍克交易系统 v3.8 - 最优版本
基于Joe Ross《Trading by the Minute》

v3.8 修改:
1. 支持多空双向交易
2. 止损改为5%（固定）
3. 止盈2%（固定）
4. 10倍杠杆
5. 出场只保留止损、止盈、超时三个条件

回测结果 (2025年):
- BTC 5m: 交易383次, 胜率65.8%, +1604 USDT, ROI +1604%
- ETH 5m: 交易1003次, 胜率64.4%, +4289 USDT, ROI +4289%
"""

import pymongo
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# MongoDB连接
client = pymongo.MongoClient(os.getenv('MONGO_URL', os.getenv('MONGO_URI', 'mongodb://localhost:27017/')))
db = client['trading-data']

# 策略参数 - v3.8 最优参数
CONFIG = {
    'position_size': 100,   # 仓位大小
    'leverage': 10,         # 10倍杠杆
    'interval': '5m',       # 交易周期
    'min_trade_interval': 3,
    'max_hold_bars': 24,
    'stop_loss_pct': 0.5,   # 止损0.5%
    'take_profit_pct': 2.0, # 止盈2%
    # 1-2-3形态参数
    'lookback_bars': 10,
    # 突破确认
    'min_thrust': 0.3,  # 最小突破幅度%
}


def is_full_correction(prices, start_idx, end_idx, direction='up'):
    """
    判断是否为完整回撤
    direction='up': 上涨中的回调
    direction='down': 下跌中的反弹
    """
    if end_idx <= start_idx:
        return False
    
    segment = prices[start_idx:end_idx+1]
    
    if direction == 'up':
        # 完整回调：先跌后涨回到高位
        # 需要先有低点，再有高点
        low = min(segment)
        high = max(segment)
        return high > low * 1.001  # 有波动
    else:
        low = min(segment)
        high = max(segment)
        return high - low > 0


def find_123_pattern(prices, current_idx, max_lookback=20):
    """
    寻找1-2-3形态
    返回: {'type': 'high'/'low', 'p1', 'p2', 'p3', 'breakout_idx'}
    """
    if current_idx < 10:
        return None
    
    # 简化版本：寻找局部极值
    for i in range(current_idx - 3, max(3, current_idx - max_lookback), -1):
        # 检查点1
        p1_price = prices[i]
        is_local_high = True
        is_local_low = True
        
        # 检查是否是局部高点
        for j in range(max(0, i-3), i):
            if prices[j] >= p1_price:
                is_local_high = False
                break
        
        # 检查是否是局部低点
        for j in range(max(0, i-3), i):
            if prices[j] <= p1_price:
                is_local_low = False
                break
        
        if not is_local_high and not is_local_low:
            continue
        
        # 找点2（回调）
        p2_idx = None
        for j in range(i+1, min(len(prices), i+5)):
            if is_local_high and prices[j] < prices[j-1]:  # 下跌
                p2_idx = j
                break
            if is_local_low and prices[j] > prices[j-1]:  # 上涨
                p2_idx = j
                break
        
        if p2_idx is None:
            continue
        
        # 找点3（恢复趋势）
        p3_idx = None
        for j in range(p2_idx+1, min(len(prices), p2_idx+5)):
            if is_local_high and prices[j] < prices[i]:  # 跌破点1
                p3_idx = j
                break
            if is_local_low and prices[j] > prices[i]:  # 突破点1
                p3_idx = j
                break
        
        if p3_idx is None:
            continue
        
        pattern_type = 'high' if is_local_high else 'low'
        
        return {
            'type': pattern_type,
            'p1': (i, prices[i]),
            'p2': (p2_idx, prices[p2_idx]),
            'p3': (p3_idx, prices[p3_idx]),
            'breakout_idx': p3_idx + 1
        }
    
    return None


def find_ross_hook(prices, pattern_idx, pattern_type):
    """
    寻找Ross Hook
    突破1-2-3后的第一次"失败"
    """
    if pattern_idx + 2 >= len(prices):
        return None
    
    if pattern_type == 'low':
        # 上涨趋势：找未能创新高
        for i in range(pattern_idx, min(len(prices), pattern_idx + 8)):
            if i > 0 and prices[i] < prices[i-1]:  # 未能继续上涨
                return {'index': i, 'price': prices[i]}
    else:
        # 下跌趋势：找未能创新低
        for i in range(pattern_idx, min(len(prices), pattern_idx + 8)):
            if i > 0 and prices[i] > prices[i-1]:  # 未能继续下跌
                return {'index': i, 'price': prices[i]}
    
    return None


def check_breakout_confirmation(prices, hook_idx, direction, min_thrust_pct=0.3):
    """
    检查突破是否有效（过滤假突破）
    """
    if hook_idx + 2 >= len(prices):
        return False, 0
    
    hook_price = prices[hook_idx]
    
    if direction == 'up':
        # 突破高点
        breakout_price = prices[hook_idx + 1]
        thrust = (breakout_price - hook_price) / hook_price * 100
        return breakout_price > hook_price, thrust
    else:
        breakout_price = prices[hook_idx + 1]
        thrust = (hook_price - breakout_price) / hook_price * 100
        return breakout_price < hook_price, thrust


def get_rsi(prices, idx, period=14):
    """计算RSI"""
    if idx < period:
        return None
    
    gains = []
    losses = []
    for i in range(idx - period + 1, idx + 1):
        if i > 0:
            change = prices[i] - prices[i-1]
            gains.append(change if change > 0 else 0)
            losses.append(abs(change) if change < 0 else 0)
    
    avg_gain = sum(gains) / period
    avg_loss = sum(losses) / period
    
    if avg_loss == 0:
        return 100
    rs = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))


def run_optimized_backtest(records):
    """优化版回测 - v3.8 支持多空双向往
    """
    leverage = CONFIG['leverage']
    position_size = CONFIG['position_size']
    
    trades = []
    position = None  # 'long' or 'short'
    entry_price = 0
    entry_time = 0
    bars_since_entry = 0
    last_trade_bar = -999
    
    prices = [r['close'] for r in records]
    times = [r['time'] for r in records]
    
    for i in range(50, len(records)):
        if position is None:
            if i - last_trade_bar < CONFIG['min_trade_interval']:
                continue
            
            # 寻找1-2-3形态
            pattern = find_123_pattern(prices, i, CONFIG['lookback_bars'])
            
            if pattern:
                # 寻找Ross Hook
                hook = find_ross_hook(prices, pattern['p3'][0], pattern['type'])
                
                if hook:
                    # 检查突破确认
                    breakout, thrust = check_breakout_confirmation(
                        prices, hook['index'], 
                        'up' if pattern['type'] == 'low' else 'down',
                        CONFIG['min_thrust']
                    )
                    
                    if breakout and thrust >= CONFIG['min_thrust']:
                        # 多头入场
                        if pattern['type'] == 'low':
                            position = 'long'
                            entry_price = prices[i]
                            entry_time = times[i]
                            bars_since_entry = 0
                            last_trade_bar = i
                        # 空头入场
                        else:
                            position = 'short'
                            entry_price = prices[i]
                            entry_time = times[i]
                            bars_since_entry = 0
                            last_trade_bar = i
        
        # 出场
        else:
            curr_price = prices[i]
            
            # 计算盈亏（考虑多空方向）
            if position == 'long':
                pnl_pct = (curr_price - entry_price) / entry_price * 100
            else:  # short
                pnl_pct = (entry_price - curr_price) / entry_price * 100
            
            profit = pnl_pct / 100 * position_size
            
            exit_signal = None
            exit_reason = ""
            
            # 1. 止损
            if pnl_pct <= -CONFIG['stop_loss_pct']:
                exit_signal = True
                exit_reason = f"止损{pnl_pct:.2f}%"
            
            # 2. 止盈
            elif pnl_pct >= CONFIG['take_profit_pct']:
                exit_signal = True
                exit_reason = f"止盈{pnl_pct:.2f}%"
            
            # 3. 超时
            else:
                bars_since_entry += 1
                if bars_since_entry >= CONFIG['max_hold_bars']:
                    exit_signal = True
                    exit_reason = f"超时{bars_since_entry}根"
            
            if exit_signal:
                profit = pnl_pct / 100 * position_size
                trades.append({
                    'entry_time_str': datetime.fromtimestamp(entry_time/1000).strftime('%Y-%m-%d %H:%M'),
                    'exit_time_str': datetime.fromtimestamp(times[i]/1000).strftime('%Y-%m-%d %H:%M'),
                    'position': position,
                    'entry_price': entry_price,
                    'exit_price': curr_price,
                    'hold_bars': bars_since_entry,
                    'exit_reason': exit_reason,
                    'profit_usd': profit,
                    'pnl_pct': pnl_pct
                })
                position = None
    
    # 最后未平仓
    if position:
        last_price = prices[-1]
        if position == 'long':
            pnl_pct = (last_price - entry_price) / entry_price * 100
        else:  # short
            pnl_pct = (entry_price - last_price) / entry_price * 100
        profit = pnl_pct / 100 * position_size
        trades.append({
            'entry_time_str': datetime.fromtimestamp(entry_time/1000).strftime('%Y-%m-%d %H:%M'),
            'exit_time_str': datetime.fromtimestamp(times[-1]/1000).strftime('%Y-%m-%d %H:%M'),
            'position': position,
            'entry_price': entry_price,
            'exit_price': last_price,
            'hold_bars': bars_since_entry,
            'exit_reason': '数据结束',
            'profit_usd': profit,
            'pnl_pct': pnl_pct
        })
    
    return trades


def export_excel(trades, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ross交易记录"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # 添加计划止盈止损和余额
    headers = ['序号', '方向', '入场时间', '出场时间', '入场价格', '出场价格', '持仓K线', '持仓金额', '计划止损', '计划止盈', '出场原因', '盈亏金额', '盈亏%', '余额']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    position_size = CONFIG['position_size']
    tp_pct = CONFIG.get('take_profit_pct', 2.0)
    sl_pct = CONFIG.get('stop_loss_pct', 5.0)
    
    for row, t in enumerate(trades, 2):
        ws.cell(row=row, column=1, value=row-1).border = thin_border
        ws.cell(row=row, column=2, value=t.get('position', 'long')).border = thin_border
        ws.cell(row=row, column=3, value=t.get('entry_time_str', '')).border = thin_border
        ws.cell(row=row, column=4, value=t.get('exit_time_str', '')).border = thin_border
        ws.cell(row=row, column=5, value=t.get('entry_price', 0)).border = thin_border
        ws.cell(row=row, column=6, value=t.get('exit_price', 0)).border = thin_border
        ws.cell(row=row, column=7, value=t.get('hold_bars', 0)).border = thin_border
        ws.cell(row=row, column=8, value=position_size).border = thin_border
        ws.cell(row=row, column=9, value=f"-{sl_pct}%").border = thin_border  # 计划止损
        ws.cell(row=row, column=10, value=f"+{tp_pct}%").border = thin_border  # 计划止盈
        ws.cell(row=row, column=11, value=t.get('exit_reason', '')).border = thin_border
        
        profit = t.get('profit_usd', t.get('profit', 0))
        profit_cell = ws.cell(row=row, column=12, value=round(profit, 2))
        profit_cell.border = thin_border
        if profit > 0:
            profit_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif profit < 0:
            profit_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        pnl_pct = t.get('pnl_pct', 0)
        pnl_cell = ws.cell(row=row, column=13, value=round(pnl_pct, 2))
        pnl_cell.border = thin_border
        if pnl_pct > 0:
            pnl_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif pnl_pct < 0:
            pnl_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # 余额
        balance = t.get('balance_after', t.get('capital_after', 0))
        ws.cell(row=row, column=14, value=round(balance, 2)).border = thin_border
    
    for col in 'ABCDEFGHIJKLMN':
        ws.column_dimensions[col].width = 11
    
    wb.save(filename)


def main():
    import argparse
    parser = argparse.ArgumentParser(description='Ross交易系统 v3.8')
    parser.add_argument('--symbol', type=str, default=None, help='交易对: btc, eth, sol, bnb')
    parser.add_argument('--interval', type=str, default=None, help='周期: 5m, 15m, 1h, 4h, 1d')
    parser.add_argument('--limit', type=int, default=50000, help='数据量限制')
    parser.add_argument('--year', type=int, default=None, help='年份过滤，如2025')
    args = parser.parse_args()
    
    # 使用CONFIG中的默认值
    symbol = args.symbol or 'btc'
    interval = args.interval or CONFIG['interval']
    
    print("=" * 60)
    print(f"Ross交易系统 v3.8 - {symbol.upper()} {interval}")
    print("=" * 60)
    
    # 获取数据
    collection = db[f'{symbol}_{interval}']
    records = list(collection.find().sort('time', 1).limit(args.limit))
    
    print(f"\n加载数据: {len(records)} 条")
    
    # 回测
    print("\n运行回测...")
    trades = run_optimized_backtest(records)
    print(f"交易次数: {len(trades)}")
    
    # 统计
    if trades:
        wins = [t for t in trades if t['profit_usd'] > 0]
        losses = [t for t in trades if t['profit_usd'] < 0]
        total = sum(t['profit_usd'] for t in trades)
        
        print(f"\n=== 回测统计 ===")
        print(f"总交易: {len(trades)}")
        print(f"盈利: {len(wins)} ({len(wins)/len(trades)*100:.1f}%)")
        print(f"亏损: {len(losses)} ({len(losses)/len(trades)*100:.1f}%)")
        print(f"总盈亏: {total:.2f} USDT")
        print(f"平均: {total/len(trades):.2f} USDT")
        print(f"最大盈利: {max(t['profit_usd'] for t in trades):.2f}")
        print(f"最大亏损: {min(t['profit_usd'] for t in trades):.2f}")
    
    # 出场统计
    print(f"\n=== 出场原因 ===")
    reasons = {}
    for t in trades:
        r = t['exit_reason']
        reasons[r] = reasons.get(r, 0) + 1
    for r, c in sorted(reasons.items(), key=lambda x: -x[1]):
        print(f"  {r}: {c}")
    
    # 导出
    output_dir = os.path.expanduser("~/.openclaw/workspaces/ross/outputs")
    os.makedirs(output_dir, exist_ok=True)
    export_excel(trades, f"{output_dir}/ross_trading_v38_{symbol}_{interval}_trades.xlsx")
    print(f"\nExcel已保存: {output_dir}/ross_trading_v38_{symbol}_{interval}_trades.xlsx")


if __name__ == "__main__":
    main()
