#!/usr/bin/env python3
"""
洛氏霍克交易系统 v2.0 - 优化版
改进：增加趋势过滤、波动率过滤、仓位管理
"""

import pymongo
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import json
import numpy as np


class StrategyConfig:
    def __init__(self):
        self.trade_amount = 100
        
        # RSI参数
        self.rsi_oversold = 30
        self.rsi_overbought = 70
        
        # 趋势判断 (EMA)
        self.ema_fast = 9    # 短期均线
        self.ema_slow = 21   # 长期均线
        
        # 波动率过滤
        self.volatility_threshold = 0.02  # 波动率阈值
        
        # 止盈止损
        self.stop_loss_pct = 0.5
        self.take_profit_pct = 1.5
        
        # 持仓限制
        self.max_hold_bars = 24
        self.min_trade_interval = 5
        
        # 信号强度要求
        self.min_signals = 2
        
    def to_dict(self):
        return {k: v for k, v in self.__dict__.items() if not k.startswith('_')}


def calculate_ema(prices, period):
    """计算EMA"""
    ema = []
    multiplier = 2 / (period + 1)
    for i, price in enumerate(prices):
        if i < period - 1:
            ema.append(None)
        elif i == period - 1:
            ema.append(sum(prices[:period]) / period)
        else:
            ema.append((price - ema[-1]) * multiplier + ema[-1])
    return ema


def calculate_volatility(prices, period=20):
    """计算波动率"""
    volatility = []
    for i in range(len(prices)):
        if i < period:
            volatility.append(None)
        else:
            returns = np.diff(prices[i-period:i]) / prices[i-period:i-1]
            volatility.append(np.std(returns))
    return volatility


class RockyHedgehogV2:
    def __init__(self, config=None):
        self.config = config or StrategyConfig()
        self.trades = []
        self.position = None
        self.entry_price = 0
        self.entry_time = 0
        self.bars_since_entry = 0
        self.last_trade_bar = -999
        
    def reset(self):
        self.trades = []
        self.position = None
        self.entry_price = 0
        self.entry_time = 0
        self.bars_since_entry = 0
        self.last_trade_bar = -999
        
    def check_entry_signal(self, curr, prev, prices, ema_fast, ema_slow, volatility, bar_index):
        """检查入场信号 - 增强版"""
        if bar_index - self.last_trade_bar < self.config.min_trade_interval:
            return None, ""
            
        rsi = curr.get('rsi')
        k = curr.get('kdj', {}).get('k')
        d = curr.get('kdj', {}).get('d')
        j = curr.get('kdj', {}).get('j')
        boll = curr.get('boll', {})
        
        prev_rsi = prev.get('rsi')
        prev_k = prev.get('kdj', {}).get('k')
        prev_d = prev.get('kdj', {}).get('d')
        prev_boll = prev.get('boll', {})
        
        if not all([rsi, k, d, boll.get('upper')]):
            return None, ""
        
        # 1. 趋势过滤：价格站在长期均线上方
        current_ema_fast = ema_fast[bar_index] if bar_index < len(ema_fast) else None
        current_ema_slow = ema_slow[bar_index] if bar_index < len(ema_slow) else None
        
        if current_ema_fast and current_ema_slow:
            trend_up = current_ema_fast > current_ema_slow
        else:
            trend_up = True
            
        # 2. 波动率过滤：波动率不能太低也不能太高
        current_vol = volatility[bar_index] if bar_index < len(volatility) and volatility[bar_index] else 0.01
        if current_vol < 0.003 or current_vol > 0.05:
            return None, "波动率不适"
            
        signals = []
        
        # RSI超卖回升
        if prev_rsi and prev_rsi < self.config.rsi_oversold and rsi > self.config.rsi_oversold:
            signals.append("RSI回升")
            
        # KDJ金叉
        if prev_k and prev_d and prev_k < prev_d and k > d:
            signals.append("KDJ金叉")
            
        # BOLL下轨支撑
        if prev_boll.get('lower') and prev['price'] < prev_boll['lower'] and curr['price'] > boll['lower']:
            signals.append("BOLL支撑")
            
        # KDJ超卖区金叉
        if k < 30 and prev_k < prev_d and k > d:
            signals.append("KDJ超卖金叉")
            
        # J值超卖反转
        if j and prev.get('kdj', {}).get('j'):
            if prev.get('kdj', {}).get('j') < 0 and j > 0:
                signals.append("J值反转")
        
        # 多信号共振
        if len(signals) >= self.config.min_signals and trend_up:
            return 'long', " + ".join(signals)
            
        return None, ""
        
    def check_exit_signal(self, curr, prev, bar_index):
        """检查出场信号"""
        if self.position != 'long':
            return None, ""
            
        rsi = curr.get('rsi')
        k = curr.get('kdj', {}).get('k')
        d = curr.get('kdj', {}).get('d')
        boll = curr.get('boll', {})
        
        prev_k = prev.get('kdj', {}).get('k')
        prev_d = prev.get('kdj', {}).get('d')
        
        pnl_pct = (curr['price'] - self.entry_price) / self.entry_price * 100
        
        # 1. 止损
        if pnl_pct <= -self.config.stop_loss_pct:
            return 'stop_loss', f"止损{abs(pnl_pct):.1f}%"
            
        # 2. 止盈
        if pnl_pct >= self.config.take_profit_pct:
            return 'take_profit', f"止盈{pnl_pct:.1f}%"
            
        # 3. KDJ死叉
        if prev_k and prev_d and prev_k > prev_d and k < d:
            return 'kdj_death', "KDJ死叉"
            
        # 4. RSI超买
        if rsi and rsi > self.config.rsi_overbought:
            return 'rsi_overbought', f"RSI超买{rsi:.0f}"
            
        # 5. 持仓时间到限
        self.bars_since_entry += 1
        if self.bars_since_entry >= self.config.max_hold_bars:
            return 'time_up', f"超时{self.bars_since_entry}根"
            
        return None, ""
        
    def run_backtest(self, records, verbose=False):
        self.reset()
        
        # 预处理数据
        prices = [r['price'] for r in records]
        ema_fast = calculate_ema(prices, self.config.ema_fast)
        ema_slow = calculate_ema(prices, self.config.ema_slow)
        volatility = calculate_volatility(prices)
        
        for i in range(1, len(records)):
            curr = records[i]
            prev = records[i-1]
            
            if self.position is None:
                signal, reason = self.check_entry_signal(
                    curr, prev, prices, ema_fast, ema_slow, volatility, i
                )
                if signal:
                    self.position = 'long'
                    self.entry_price = curr['price']
                    self.entry_time = curr['time']
                    self.bars_since_entry = 0
                    self.last_trade_bar = i
                    
            else:
                exit_signal, exit_reason = self.check_exit_signal(curr, prev, i)
                if exit_signal:
                    profit = (curr['price'] - self.entry_price) / self.entry_price * self.config.trade_amount
                    pnl_pct = (curr['price'] - self.entry_price) / self.entry_price * 100
                    
                    self.trades.append({
                        'entry_time': datetime.fromtimestamp(self.entry_time/1000).strftime('%Y-%m-%d %H:%M'),
                        'exit_time': datetime.fromtimestamp(curr['time']/1000).strftime('%Y-%m-%d %H:%M'),
                        'entry_price': self.entry_price,
                        'exit_price': curr['price'],
                        'hold_bars': self.bars_since_entry,
                        'exit_reason': exit_reason,
                        'profit_usd': profit,
                        'pnl_pct': pnl_pct
                    })
                    self.position = None
                    
        # 平仓
        if self.position == 'long':
            last = records[-1]
            profit = (last['price'] - self.entry_price) / self.entry_price * self.config.trade_amount
            pnl_pct = (last['price'] - self.entry_price) / self.entry_price * 100
            self.trades.append({
                'entry_time': datetime.fromtimestamp(self.entry_time/1000).strftime('%Y-%m-%d %H:%M'),
                'exit_time': datetime.fromtimestamp(last['time']/1000).strftime('%Y-%m-%d %H:%M'),
                'entry_price': self.entry_price,
                'exit_price': last['price'],
                'hold_bars': self.bars_since_entry,
                'exit_reason': '数据结束',
                'profit_usd': profit,
                'pnl_pct': pnl_pct
            })
            
        return self.get_stats()
        
    def get_stats(self):
        if not self.trades:
            return {
                'trade_count': 0, 'win_rate': 0, 'avg_profit': 0,
                'avg_loss': 0, 'profit_factor': 0, 'total_pnl': 0,
                'max_consecutive_wins': 0, 'max_consecutive_losses': 0
            }
            
        wins = [t for t in self.trades if t['profit_usd'] > 0]
        losses = [t for t in self.trades if t['profit_usd'] <= 0]
        
        win_rate = len(wins) / len(self.trades) * 100
        avg_profit = sum(t['profit_usd'] for t in wins) / len(wins) if wins else 0
        avg_loss = sum(t['profit_usd'] for t in losses) / len(losses) if losses else 0
        
        total_win = sum(t['profit_usd'] for t in wins)
        total_loss = abs(sum(t['profit_usd'] for t in losses))
        profit_factor = total_win / total_loss if total_loss > 0 else float('inf')
        
        # 连胜连负
        max_wins = 0
        max_losses = 0
        cw = 0
        cl = 0
        for t in self.trades:
            if t['profit_usd'] > 0:
                cw += 1
                cl = 0
                max_wins = max(max_wins, cw)
            else:
                cl += 1
                cw = 0
                max_losses = max(max_losses, cl)
                
        return {
            'trade_count': len(self.trades),
            'win_rate': win_rate,
            'avg_profit': avg_profit,
            'avg_loss': avg_loss,
            'profit_factor': profit_factor,
            'total_pnl': sum(t['profit_usd'] for t in self.trades),
            'max_consecutive_wins': max_wins,
            'max_consecutive_losses': max_losses
        }


def load_data(limit=2000):
    client = pymongo.MongoClient('mongodb://localhost:27017/')
    db = client['trading-data']
    collection = db['btc_5m']
    records = list(collection.find().sort('time', 1).limit(limit))
    for r in records:
        r.pop('_id', None)
    return records


def optimize_v2(records):
    """参数优化 v2"""
    print("\n" + "="*60)
    print("洛氏霍克交易系统 v2.0 优化")
    print("="*60)
    
    best_score = -float('inf')
    best_config = None
    best_stats = None
    
    results = []
    
    # 参数网格 - 扩大范围
    for rsi_oversold in [25, 30, 35]:
        for rsi_overbought in [65, 70, 75, 80]:
            if rsi_oversold >= rsi_overbought:
                continue
            for ema_fast in [7, 9, 12]:
                for ema_slow in [21, 26, 30]:
                    if ema_fast >= ema_slow:
                        continue
                    for stop_loss in [0.3, 0.5, 0.7]:
                        for take_profit in [1.0, 1.2, 1.5, 2.0]:
                            if take_profit <= stop_loss:
                                continue
                            for max_hold in [12, 24, 36]:
                                for min_interval in [3, 5, 7]:
                                    for min_signals in [2, 3]:
                                        config = StrategyConfig()
                                        config.rsi_oversold = rsi_oversold
                                        config.rsi_overbought = rsi_overbought
                                        config.ema_fast = ema_fast
                                        config.ema_slow = ema_slow
                                        config.stop_loss_pct = stop_loss
                                        config.take_profit_pct = take_profit
                                        config.max_hold_bars = max_hold
                                        config.min_trade_interval = min_interval
                                        config.min_signals = min_signals
                                        
                                        system = RockyHedgehogV2(config)
                                        stats = system.run_backtest(records, verbose=False)
                                        
                                        if stats['trade_count'] < 3:
                                            continue
                                        
                                        # 评分
                                        # 胜率权重35
                                        win_score = min(stats['win_rate'] / 100, 1) * 35
                                        
                                        # 盈亏比权重35
                                        pf = stats['profit_factor'] if stats['profit_factor'] != float('inf') else 5
                                        pf_score = min(pf / 3, 1) * 35
                                        
                                        # 交易频率权重30 (每天1-4笔)
                                        days = 3.5
                                        tpd = stats['trade_count'] / days
                                        if 1 <= tpd <= 4:
                                            freq_score = 30
                                        elif tpd < 1:
                                            freq_score = tpd * 20
                                        else:
                                            freq_score = max(0, 30 - (tpd - 4) * 8)
                                        
                                        total_score = win_score + pf_score + freq_score
                                        
                                        results.append({
                                            'config': config,
                                            'stats': stats,
                                            'score': total_score
                                        })
                                        
                                        if total_score > best_score:
                                            best_score = total_score
                                            best_config = config
                                            best_stats = stats
    
    results.sort(key=lambda x: x['score'], reverse=True)
    
    print("\n=== TOP 10 优化结果 ===")
    for i, r in enumerate(results[:10], 1):
        s = r['stats']
        c = r['config']
        print(f"\n#{i} 评分:{r['score']:.1f} | 交易:{s['trade_count']} | "
              f"胜率:{s['win_rate']:.1f}% | 盈亏比:{s['profit_factor']:.2f} | 总盈亏:${s['total_pnl']:.2f}")
        print(f"   RSI:{c.rsi_oversold}/{c.rsi_overbought} EMA:{c.ema_fast}/{c.ema_slow} "
              f"止盈:{c.take_profit_pct}% 止损:{c.stop_loss_pct}%")
    
    return best_config, best_stats, results[:10]


def run_final(records, config):
    print("\n" + "="*60)
    print("最终测试")
    print("="*60)
    
    system = RockyHedgehogV2(config)
    stats = system.run_backtest(records)
    
    print(f"\n交易次数: {stats['trade_count']}")
    print(f"胜率: {stats['win_rate']:.1f}%")
    print(f"平均盈利: ${stats['avg_profit']:.2f}")
    print(f"平均亏损: ${stats['avg_loss']:.2f}")
    print(f"盈亏比: {stats['profit_factor']:.2f}")
    print(f"总盈亏: ${stats['total_pnl']:.2f}")
    print(f"最大连胜: {stats['max_consecutive_wins']}")
    print(f"最大连亏: {stats['max_consecutive_losses']}")
    
    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "交易记录"
    
    headers = ['入场时间', '出场时间', '入场价', '出场价', '持仓', '出场理由', '盈亏($)', '盈亏(%)']
    hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hf_font = Font(bold=True, color="FFFFFF")
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hf
        cell.font = hf_font
    
    for row, t in enumerate(system.trades, 2):
        ws.cell(row=row, column=1, value=t['entry_time'])
        ws.cell(row=row, column=2, value=t['exit_time'])
        ws.cell(row=row, column=3, value=t['entry_price'])
        ws.cell(row=row, column=4, value=t['exit_price'])
        ws.cell(row=row, column=5, value=t['hold_bars'])
        ws.cell(row=row, column=6, value=t['exit_reason'])
        ws.cell(row=row, column=7, value=round(t['profit_usd'], 2))
        ws.cell(row=row, column=8, value=f"{t['pnl_pct']:.2f}%")
        
        pc = ws.cell(row=row, column=7)
        if t['profit_usd'] >= 0:
            pc.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            pc.font = Font(color="006100")
        else:
            pc.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            pc.font = Font(color="9C0006")
    
    sr = len(system.trades) + 3
    ws.cell(row=sr, column=1, value="总计").font = Font(bold=True)
    ws.cell(row=sr, column=7, value=f"${stats['total_pnl']:.2f}")
    
    for col in 'ABCDEFGH':
        ws.column_dimensions[col].width = 16
    
    output = '/Users/makaihong/.openclaw/workspaces/ross/洛氏霍克交易系统_v2.xlsx'
    wb.save(output)
    print(f"\n已保存: {output}")
    
    return stats


if __name__ == "__main__":
    records = load_data(2000)
    print(f"数据: {len(records)}条 | {datetime.fromtimestamp(records[0]['time']/1000)} ~ {datetime.fromtimestamp(records[-1]['time']/1000)}")
    
    best_config, best_stats, top_results = optimize_v2(records)
    stats = run_final(records, best_config)
    
    # 保存配置
    with open('/Users/makaihong/.openclaw/workspaces/ross/trading_config_v2.json', 'w') as f:
        json.dump({
            'config': best_config.to_dict(),
            'stats': best_stats
        }, f, indent=2)
    print("配置已保存")
